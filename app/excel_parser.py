"""Parse a BSM1 Excel template into PlantParameters + SCADA tag metadata."""

import contextlib
import io

from openpyxl import load_workbook

from .models import PlantParameters, SimulationRequest

# Mapa: api_id del Excel -> campo real de PlantParameters (solo los que difieren)
_ID_MAP = {
    'V_anoxic_A1': 'V_anoxic',
    'V_anoxic_A2': '_aux_V_anoxic',
    'V_anoxic_A3': '_skip',
    'V_aerobic_O1': 'V_aerobic',
    'V_aerobic_O2': '_aux_V_aerobic',
    'V_aerobic_O3': '_aux_V_aerobic',
    'V_aerobic_O4': '_skip',
    'Q2': '_skip',
    'COD_total2': '_skip',
    'TKN2': '_skip',
    'TSS_inf2': '_skip',
}

_FLOAT_PARAMS = {
    'Q',
    'COD_total',
    'TKN',
    'TSS_inf',
    'V_anoxic',
    'V_aerobic',
    'Clarifier_area',
    'Clarifier_height',
    'Q_RAS',
    'Q_WAS',
    'SRT_target',
    'Temperature',
    'DO_O1',
    'DO_O2',
    'DO_O3',
    'KLa_O1',
    'KLa_O2',
    'KLa_O3',
    'DOsat',
    'Q_intr',
}
_STR_PARAMS = {'sludge_control', 'aeration_control'}
_SIM_PARAMS: dict[str, type] = {'t_days': int, 'method': str, 'tolerance': float}

# Claves auxiliares para comparacion de volumenes por reactor
_AUX_KEYS = {'_aux_V_anoxic', '_aux_V_aerobic'}


def _check_volume_warnings(
    plant_data: dict[str, object],
    aux_data: dict[str, list[float]],
) -> list[str]:
    """Compare per-reactor volumes and warn if they differ >1%."""
    warnings: list[str] = []
    checks = [
        ('V_anoxic', '_aux_V_anoxic', 'anoxico', 'A1', 'A2'),
        ('V_aerobic', '_aux_V_aerobic', 'aerobico', 'O1', 'O2/O3'),
    ]
    for field, aux_key, label, primary, secondary in checks:
        primary_val = plant_data.get(field)
        aux_vals = aux_data.get(aux_key, [])
        if primary_val is None or not aux_vals:
            continue
        for v in aux_vals:
            if abs(v - primary_val) / max(primary_val, 1e-6) > 0.01:
                warnings.append(
                    f'Volumenes reactor {label} difieren: '
                    f'{primary}={primary_val} m3, {secondary}={v} m3. '
                    f'Se usa {primary}={primary_val} m3 para todos.'
                )
                break
    return warnings


def parse_bsm1_excel(
    file_bytes: bytes,
) -> tuple[SimulationRequest, dict[str, str], list[str]]:
    """Parse uploaded Excel -> (SimulationRequest, tags_dict, warnings).

    tags_dict maps param ID -> 'tag_value|source' string.
    warnings contains messages about ignored or inconsistent data.
    Fields missing or empty in Excel use PlantParameters defaults.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb['Par\u00e1metros_BSM1']

        plant_data: dict[str, object] = {}
        sim_data: dict[str, object] = {}
        tags: dict[str, str] = {}
        aux_data: dict[str, list[float]] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            api_id = row[1]  # col B: ID API
            value = row[2]  # col C: Valor
            tag = row[6]  # col G: Tag
            source = row[7]  # col H: Fuente

            if not api_id or not str(api_id).strip():
                continue  # fila de seccion header o vacia

            api_id = str(api_id).strip()

            # Registrar tag si existe
            if tag and str(tag).strip():
                tags[api_id] = f'{str(tag).strip()}|{source or "Manual"}'

            if value is None:
                continue

            # Aplicar mapeo de IDs
            mapped = _ID_MAP.get(api_id, api_id)
            if mapped == '_skip':
                continue

            # Valores auxiliares para comparacion de volumenes
            if mapped in _AUX_KEYS:
                with contextlib.suppress(ValueError, TypeError):
                    aux_data.setdefault(mapped, []).append(float(value))
                continue

            if mapped in _FLOAT_PARAMS:
                with contextlib.suppress(ValueError, TypeError):
                    plant_data[mapped] = float(value)
            elif mapped in _STR_PARAMS:
                plant_data[mapped] = str(value).strip()
            elif mapped in _SIM_PARAMS:
                with contextlib.suppress(ValueError, TypeError):
                    sim_data[mapped] = _SIM_PARAMS[mapped](value)
    finally:
        wb.close()

    # Comparar volumenes por reactor y generar warnings
    excel_warnings = _check_volume_warnings(plant_data, aux_data)

    # Construir PlantParameters (Pydantic rellena campos faltantes con defaults)
    plant = PlantParameters(**plant_data)
    request = SimulationRequest(plant=plant, **sim_data)
    return request, tags, excel_warnings
