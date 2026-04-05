"""Generate a BSM1 parameter template as an Excel (.xlsx) file in memory."""

import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Row definition
# ---------------------------------------------------------------------------


@dataclass
class _Row:
    label: str
    api_id: str
    valor: float | int | str | None
    unidad: str
    min_val: float | int | None
    max_val: float | int | None
    editable: bool = True
    range_override: str | None = None
    dropdown: str | None = None  # comma-separated options for col C


# ---------------------------------------------------------------------------
# Section / row data
# ---------------------------------------------------------------------------

_SECTIONS: list[tuple[str, list[_Row]]] = [
    (
        'INFLUENTE PRINCIPAL',
        [
            _Row('Caudal influente', 'Q', 18446, 'm\u00b3/d', 1000, 200000),
            _Row('COD total', 'COD_total', 381.19, 'mg/L', 50, 2000),
            _Row('TKN influente', 'TKN', 51.35, 'mg N/L', 10, 300),
            _Row('TSS influente', 'TSS_inf', 211.27, 'mg/L', 20, 1500),
        ],
    ),
    (
        'SEGUNDO INFLUENTE (reservado v2.0 \u2014 no soportado en simulaci\u00f3n)',
        [
            _Row(
                'Caudal influente 2 (*)',
                'Q2',
                None,
                'm\u00b3/d',
                1000,
                200000,
                range_override='No soportado v1.0',
            ),
            _Row(
                'COD total 2 (*)',
                'COD_total2',
                None,
                'mg/L',
                50,
                2000,
                range_override='No soportado v1.0',
            ),
            _Row(
                'TKN influente 2 (*)',
                'TKN2',
                None,
                'mg N/L',
                10,
                300,
                range_override='No soportado v1.0',
            ),
            _Row(
                'TSS influente 2 (*)',
                'TSS_inf2',
                None,
                'mg/L',
                20,
                1500,
                range_override='No soportado v1.0',
            ),
        ],
    ),
    (
        'GEOMETR\u00cdA \u2014 REACTORES AN\u00d3XICOS',
        [
            _Row('Vol. an\u00f3xico A1', 'V_anoxic_A1', 1000, 'm\u00b3', 100, 20000),
            _Row('Vol. an\u00f3xico A2', 'V_anoxic_A2', 1000, 'm\u00b3', 100, 20000),
            _Row(
                'Vol. an\u00f3xico A3 (*)',
                'V_anoxic_A3',
                None,
                'm\u00b3',
                100,
                20000,
                range_override='No soportado v1.0',
            ),
        ],
    ),
    (
        'GEOMETR\u00cdA \u2014 REACTORES AER\u00d3BICOS',
        [
            _Row('Vol. aer\u00f3bico O1', 'V_aerobic_O1', 1333, 'm\u00b3', 100, 20000),
            _Row('Vol. aer\u00f3bico O2', 'V_aerobic_O2', 1333, 'm\u00b3', 100, 20000),
            _Row('Vol. aer\u00f3bico O3', 'V_aerobic_O3', 1333, 'm\u00b3', 100, 20000),
            _Row(
                'Vol. aer\u00f3bico O4 (*)',
                'V_aerobic_O4',
                None,
                'm\u00b3',
                100,
                20000,
                range_override='No soportado v1.0',
            ),
        ],
    ),
    (
        'GEOMETR\u00cdA \u2014 DECANTADOR SECUNDARIO',
        [
            _Row('\u00c1rea decantador', 'Clarifier_area', 1500, 'm\u00b2', 100, 10000),
            _Row('Altura decantador', 'Clarifier_height', 4, 'm', 1, 10),
        ],
    ),
    (
        'OPERACI\u00d3N \u2014 CONTROL DE FANGOS',
        [
            _Row(
                'Modo control fangos',
                'sludge_control',
                'WAS',
                'WAS/SRT',
                None,
                None,
                dropdown='WAS,SRT',
            ),
            _Row('Purga (WAS)', 'Q_WAS', 385, 'm\u00b3/d', 10, 5000),
            _Row('SRT objetivo', 'SRT_target', None, 'd\u00edas', 1, 40),
        ],
    ),
    (
        'OPERACI\u00d3N \u2014 RECIRCULACIONES',
        [
            _Row('Recirculaci\u00f3n externa (RAS)', 'Q_RAS', 18446, 'm\u00b3/d', 0, 200000),
            _Row('Recirculaci\u00f3n interna', 'Q_intr', 55338, 'm\u00b3/d', 0, 300000),
        ],
    ),
    (
        'OPERACI\u00d3N \u2014 TEMPERATURA',
        [
            _Row('Temperatura', 'Temperature', 20, '\u00b0C', 5, 35),
        ],
    ),
    (
        'AIREACI\u00d3N',
        [
            _Row(
                'Modo aireaci\u00f3n',
                'aeration_control',
                'DO',
                'DO/KLa',
                None,
                None,
                dropdown='DO,KLa',
            ),
            _Row('Consigna DO reactor O1', 'DO_O1', 1.72, 'mg O2/L', 0.0, 7.0),
            _Row('Consigna DO reactor O2', 'DO_O2', 2.43, 'mg O2/L', 0.0, 7.0),
            _Row('Consigna DO reactor O3', 'DO_O3', 0.49, 'mg O2/L', 0.0, 7.0),
            _Row('KLa reactor O1 (modo KLa)', 'KLa_O1', 240, '1/d', 0, 1000),
            _Row('KLa reactor O2 (modo KLa)', 'KLa_O2', 240, '1/d', 0, 1000),
            _Row('KLa reactor O3 (modo KLa)', 'KLa_O3', 84, '1/d', 0, 1000),
            _Row('OD saturaci\u00f3n', 'DOsat', 8.0, 'mg O2/L', 6.0, 14.0),
        ],
    ),
    (
        'PAR\u00c1METROS DE SIMULACI\u00d3N',
        [
            _Row('D\u00edas de simulaci\u00f3n', 't_days', 200, 'd\u00edas', 50, 1000),
            _Row('M\u00e9todo ODE', 'method', 'BDF', 'BDF/Radau', None, None, dropdown='BDF,Radau'),
            _Row('Tolerancia validaci\u00f3n', 'tolerance', 0.05, 'fracci\u00f3n', 0.01, 0.5),
        ],
    ),
]


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_FILL_GREY_LIGHT = PatternFill('solid', fgColor='ECEFF1')
_FILL_GREY_DARK = PatternFill('solid', fgColor='CFD8DC')
_FILL_YELLOW = PatternFill('solid', fgColor='FFFDE7')
_FILL_BLUE = PatternFill('solid', fgColor='E3F2FD')
_FILL_HEADER = PatternFill('solid', fgColor='37474F')
_FONT_HEADER = Font(bold=True, color='FFFFFF', size=11)
_FONT_SECTION = Font(bold=True, color='FFFFFF', size=11)
_FONT_DEFAULT = Font(size=11)
_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

_COL_WIDTHS = {1: 32, 2: 18, 3: 13, 4: 13, 5: 13, 6: 20, 7: 30, 8: 18}

_HEADERS = [
    'Par\u00e1metro',
    'ID (API)',
    'Valor',
    'Unidad',
    'Ref. BSM1',
    'Rango',
    'Tag / ID variable',
    'Fuente',
]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_bsm1_template() -> bytes:
    """Return .xlsx template as bytes."""
    wb = Workbook()

    # ---- Sheet 1: Parametros_BSM1 -----------------------------------------
    ws = wb.active
    ws.title = 'Par\u00e1metros_BSM1'

    # Column widths
    for col_num, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Header row (row 1)
    for col_idx, header_text in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER

    # DataValidation for column H (Fuente) — applied to all data rows later
    dv_fuente = DataValidation(
        type='list',
        formula1='"SCADA,Laboratorio,Informe,Manual"',
        allow_blank=True,
    )
    dv_fuente.error = 'Selecciona una fuente v\u00e1lida.'
    dv_fuente.errorTitle = 'Fuente no v\u00e1lida'
    ws.add_data_validation(dv_fuente)

    current_row = 2

    for section_title, rows in _SECTIONS:
        # Section header row (merged A:H)
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=8,
        )
        sec_cell = ws.cell(row=current_row, column=1, value=section_title)
        sec_cell.fill = _FILL_HEADER
        sec_cell.font = _FONT_SECTION
        sec_cell.alignment = Alignment(
            horizontal='left',
            vertical='center',
            indent=1,
        )
        current_row += 1

        for r in rows:
            # Col A — label
            ca = ws.cell(row=current_row, column=1, value=r.label)
            ca.fill = _FILL_GREY_LIGHT
            ca.font = _FONT_DEFAULT
            ca.alignment = _ALIGN_LEFT

            # Col B — API id
            cb = ws.cell(row=current_row, column=2, value=r.api_id)
            cb.fill = _FILL_GREY_LIGHT
            cb.font = _FONT_DEFAULT
            cb.alignment = _ALIGN_CENTER

            # Col C — value (editable, yellow)
            cc = ws.cell(row=current_row, column=3, value=r.valor)
            cc.fill = _FILL_YELLOW
            cc.font = _FONT_DEFAULT
            cc.alignment = _ALIGN_CENTER

            # Dropdown on C if specified
            if r.dropdown:
                dv_c = DataValidation(
                    type='list',
                    formula1=f'"{r.dropdown}"',
                    allow_blank=True,
                )
                ws.add_data_validation(dv_c)
                dv_c.add(cc)

            # Col D — unit
            cd = ws.cell(row=current_row, column=4, value=r.unidad)
            cd.fill = _FILL_GREY_LIGHT
            cd.font = _FONT_DEFAULT
            cd.alignment = _ALIGN_CENTER

            # Col E — BSM1 reference (same as default value)
            ce = ws.cell(row=current_row, column=5, value=r.valor)
            ce.fill = _FILL_GREY_DARK
            ce.font = _FONT_DEFAULT
            ce.alignment = _ALIGN_CENTER

            # Col F — range
            if r.range_override:
                range_text = r.range_override
            elif r.min_val is not None and r.max_val is not None:
                range_text = f'{r.min_val} \u2013 {r.max_val}'
            else:
                range_text = ''
            cf = ws.cell(row=current_row, column=6, value=range_text)
            cf.fill = _FILL_GREY_LIGHT
            cf.font = _FONT_DEFAULT
            cf.alignment = _ALIGN_CENTER

            # Col G — Tag / ID variable (editable, blue)
            cg = ws.cell(row=current_row, column=7)
            cg.fill = _FILL_BLUE
            cg.font = _FONT_DEFAULT
            cg.alignment = _ALIGN_LEFT

            # Col H — Fuente (editable, blue, dropdown)
            ch = ws.cell(row=current_row, column=8)
            ch.fill = _FILL_BLUE
            ch.font = _FONT_DEFAULT
            ch.alignment = _ALIGN_CENTER
            dv_fuente.add(ch)

            current_row += 1

    # ---- Sheet 2: Instrucciones -------------------------------------------
    ws2 = wb.create_sheet('Instrucciones')
    ws2.column_dimensions['A'].width = 90

    instructions = [
        'Instrucciones de uso \u2014 Plantilla BSM1',
        '',
        '1. Celdas amarillas (columna C): introduce el valor de la planta. '
        'Si se deja vac\u00edo, se usa el default BSM1.',
        '2. Celdas azules (columna G): introduce el identificador de la '
        'variable en tu sistema SCADA, laboratorio o informe.',
        '3. Celdas azules (columna H): selecciona la fuente del dato '
        '(SCADA / Laboratorio / Informe / Manual).',
        '4. Columna E: valor de referencia BSM1 (no editar).',
        '5. Filas marcadas con (*): no soportadas en v1.0, ser\u00e1n '
        'ignoradas en la simulaci\u00f3n.',
        '6. Guardado: guarda como .xlsx (no .xls ni .xlsm).',
        '7. Carga: usa el bot\u00f3n "Cargar y simular" en la interfaz web del simulador.',
    ]

    title_font = Font(bold=True, size=14)
    for idx, line in enumerate(instructions, start=1):
        cell = ws2.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = title_font

    # ---- Write to bytes ---------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
