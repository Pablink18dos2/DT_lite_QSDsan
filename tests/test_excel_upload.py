"""Test: fill BSM1 Excel template with a 10 MLD high-industrial-load plant and upload."""

import json
import sys

import requests
from openpyxl import load_workbook

API = 'http://localhost:8000'
TEMPLATE_PATH = 'test_template.xlsx'
FILLED_PATH = 'test_10MLD_industrial.xlsx'

# ============================================================================
# PLANT: 10 MLD with high industrial load
# ============================================================================
# Justification:
#   Q = 10 000 m3/d (10 MLD)
#   COD = 900 mg/L (alta carga industrial: conservera, lactea, etc.)
#   TKN = 65 mg N/L (algo elevado por vertidos industriales)
#   TSS = 450 mg/L (alta carga de solidos)
#   C:N ratio = 900/65 = 13.8 (OK para desnitrificacion)
#   Reactores dimensionados para 10 MLD:
#     V_anoxic = 700 m3 cada uno (TRH_anox ~3.4h para Q+RAS+intr)
#     V_aerobic = 1100 m3 cada uno (TRH_aer ~7.9h para Q+RAS)
#   Clarificador: 800 m2 (SOR = 10000/800 = 0.52 m/h, conservador)
#   RAS = 1Q, Q_intr = 3Q
#   Q_WAS = 200 m3/d (SRT ~12-15 d)
#   T = 18 C (invierno suave, planta real)
#   DO elevado en O1/O2 por alta carga COD: 3.0 / 3.5 mg O2/L
#   DO_O3 = 0.5 (proteger zona anoxica)
#   DOsat = 8.5 mg O2/L (a 18C, ~20m altitud)
# ============================================================================

VALUES = {
    'Q': 10000,
    'COD_total': 900,
    'TKN': 65,
    'TSS_inf': 450,
    'V_anoxic_A1': 700,
    'V_anoxic_A2': 700,
    'V_aerobic_O1': 1100,
    'V_aerobic_O2': 1100,
    'V_aerobic_O3': 1100,
    'Clarifier_area': 800,
    'Clarifier_height': 4,
    'sludge_control': 'WAS',
    'Q_WAS': 200,
    'Q_RAS': 10000,
    'Q_intr': 30000,
    'Temperature': 18,
    'aeration_control': 'DO',
    'DO_O1': 3.0,
    'DO_O2': 3.5,
    'DO_O3': 0.5,
    'DOsat': 8.5,
    't_days': 200,
    'method': 'BDF',
    'tolerance': 0.05,
}

TAGS = {
    'Q': ('FT-001', 'SCADA'),
    'COD_total': ('AQ-COD-ENT', 'Laboratorio'),
    'TKN': ('AQ-TKN-ENT', 'Laboratorio'),
    'TSS_inf': ('AQ-SST-ENT', 'Laboratorio'),
    'Temperature': ('TT-BIO-01', 'SCADA'),
    'DO_O1': ('DO-O1-SP', 'SCADA'),
    'DO_O2': ('DO-O2-SP', 'SCADA'),
    'DO_O3': ('DO-O3-SP', 'SCADA'),
    'Q_RAS': ('FT-RAS-01', 'SCADA'),
    'Q_WAS': ('FT-PUR-01', 'SCADA'),
}


def fill_template():
    """Fill the Excel template with plant values and SCADA tags."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['Par\u00e1metros_BSM1']

    filled = 0
    for row in ws.iter_rows(min_row=2):
        api_id_cell = row[1]  # col B
        val_cell = row[2]  # col C
        tag_cell = row[6]  # col G
        src_cell = row[7]  # col H

        api_id = api_id_cell.value
        if not api_id:
            continue
        api_id = str(api_id).strip()

        if api_id in VALUES:
            val_cell.value = VALUES[api_id]
            filled += 1

        if api_id in TAGS:
            tag_cell.value = TAGS[api_id][0]
            src_cell.value = TAGS[api_id][1]

    wb.save(FILLED_PATH)
    print(f'Filled {filled} parameters, saved to {FILLED_PATH}')
    return FILLED_PATH


def upload_and_simulate(filepath):
    """Upload the filled Excel and run the simulation."""
    with open(filepath, 'rb') as f:
        resp = requests.post(
            f'{API}/simulate/from-excel',
            files={
                'file': (
                    'BSM1_10MLD.xlsx',
                    f,
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
            },
            timeout=120,
        )

    if resp.status_code != 200:
        print(f'ERROR HTTP {resp.status_code}: {resp.text}')
        sys.exit(1)

    return resp.json()


def analyze_results(result):
    """Print key results and process engineering analysis."""
    agg = result['aggregates']
    comps = {c['component']: c for c in result['components']}

    print('\n' + '=' * 70)
    print('RESULTS: 10 MLD PLANT WITH HIGH INDUSTRIAL LOAD')
    print('=' * 70)

    print('\n--- Simulation ---')
    print(f'  Method: {result["method_used"]}')
    print(f'  Compute time: {result["compute_time_seconds"]:.1f} s')
    print(f'  Validation: {result["pass_count"]}/{result["total_count"]} PASS')

    print('\n--- Key Effluent Quality ---')
    keys = ['S_NH', 'S_NO', 'S_O', 'S_S', 'X_BH', 'X_BA']
    for k in keys:
        if k in comps:
            c = comps[k]
            status = 'PASS' if c['passed'] else 'FAIL'
            print(
                f'  {k:8s}: {c["obtained"]:8.3f} {c["unit"]:15s} '
                f'(ref: {c["reference"]:.3f}, err: {c["error_pct"]:+.1f}%) [{status}]'
            )

    print('\n--- Aggregated Parameters ---')
    agg_keys = [
        ('COD_eff', 'mg COD/L'),
        ('TN_eff', 'mg N/L'),
        ('TSS_eff', 'mg SS/L'),
        ('BOD5_eff', 'mg/L'),
        ('SRT_traditional', 'days'),
        ('SRT_specific', 'days'),
        ('HRT_total', 'h'),
        ('Q_eff', 'm3/d'),
    ]
    for k, unit in agg_keys:
        v = agg.get(k)
        if v is not None:
            print(f'  {k:20s}: {v:10.3f} {unit}')

    # Removal percentages
    print('\n--- Removal Efficiency ---')
    rem_keys = ['removal_COD_pct', 'removal_TN_pct', 'removal_NH4_pct', 'removal_TSS_pct']
    for k in rem_keys:
        v = agg.get(k)
        if v is not None:
            print(f'  {k:22s}: {v:.1f}%')

    # SCADA tags echoed
    print('\n--- SCADA Tags (echoed) ---')
    tag_count = 0
    for k, v in sorted(agg.items()):
        if k.startswith('_tag_'):
            print(f'  {k:25s}: {v}')
            tag_count += 1
    print(f'  Total tags echoed: {tag_count}')

    # Applied parameters
    print('\n--- Applied Parameters ---')
    param_keys = [
        '_param_Q',
        '_param_COD_total',
        '_param_TKN',
        '_param_TSS_inf',
        '_param_V_anoxic',
        '_param_V_aerobic',
        '_param_Temperature',
        '_param_Q_WAS',
        '_param_Q_RAS',
        '_param_Q_intr',
        '_param_DO_O1',
        '_param_DO_O2',
        '_param_DO_O3',
        '_param_DOsat',
        '_param_KLa_O1',
        '_param_KLa_O2',
        '_param_KLa_O3',
    ]
    for k in param_keys:
        v = agg.get(k)
        if v is not None:
            print(f'  {k:25s}: {v}')

    # IWA Limits check
    print('\n--- IWA Effluent Limits Check ---')
    limits = {'COD_eff': 100, 'BOD5_eff': 10, 'TN_eff': 18, 'TSS_eff': 30}
    snh = agg.get('S_NH_eff') or (comps['S_NH']['obtained'] if 'S_NH' in comps else None)
    if snh is not None:
        limits['S_NH'] = 4
        agg['S_NH'] = snh

    for param, limit in limits.items():
        v = agg.get(param)
        if v is not None:
            status = 'OK' if v < limit else '!! VIOLATION'
            print(f'  {param:10s}: {v:8.3f} (limit {limit}) [{status}]')

    print('\n' + '=' * 70)


if __name__ == '__main__':
    path = fill_template()
    result = upload_and_simulate(path)
    analyze_results(result)

    # Also save full JSON for inspection
    with open('test_10MLD_result.json', 'w') as f:
        json.dump(result, f, indent=2)
    print('\nFull result saved to test_10MLD_result.json')
